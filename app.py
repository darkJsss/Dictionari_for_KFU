from flask import Flask, render_template, request, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(os.path.abspath(os.path.dirname(__file__)),
                                                                    'data.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join('templates', 'excel')

db = SQLAlchemy(app)


class Suggestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    word = db.Column(db.String(100), nullable=False)
    russian_equivalent = db.Column(db.String(100))
    origin = db.Column(db.String(200))
    comment = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='pending')


# Создаем таблицы при запуске приложения
def create_tables():
    with app.app_context():
        db.create_all()
        print("Таблицы успешно созданы")


create_tables()


def load_dictionary():
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'dictionary1.xlsx')
    words = []

    try:
        wb = load_workbook(excel_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                word = {
                    'Название': str(row[0]),
                    'Русский аналог': str(row[1]) if len(row) > 1 and row[1] else '',
                    'Происхождение': str(row[2]) if len(row) > 2 and row[2] else ''
                }
                words.append(word)

    except Exception as e:
        print(f"Ошибка загрузки Excel файла: {e}")

    return words


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/dictionary')
def dictionary():
    search_term = request.args.get('search', '').lower()
    letter_filter = request.args.get('letter', '')
    words = load_dictionary()

    russian_alphabet = ['А', 'Б', 'В', 'Г', 'Д', 'Е', 'Ё', 'Ж', 'З', 'И', 'Й',
                        'К', 'Л', 'М', 'Н', 'О', 'П', 'Р', 'С', 'Т', 'У', 'Ф',
                        'Х', 'Ц', 'Ч', 'Ш', 'Щ', 'Ъ', 'Ы', 'Ь', 'Э', 'Ю', 'Я']

    if letter_filter:
        filtered_words = [word for word in words if word['Название'].upper().startswith(letter_filter)]
    else:
        filtered_words = words

    if search_term:
        filtered_words = [word for word in filtered_words if search_term in word['Название'].lower()]

    return render_template(
        'dictionary.html',
        words=filtered_words,
        search_term=search_term,
        letter_filter=letter_filter,
        alphabet=russian_alphabet
    )


@app.route('/suggest', methods=['GET', 'POST'])
def suggest():
    if request.method == 'POST':
        word = request.form.get('word', '').strip()
        russian_equivalent = request.form.get('russian_equivalent', '').strip()
        origin = request.form.get('origin', '').strip()
        comment = request.form.get('comment', '').strip()

        if word:
            try:
                with app.app_context():
                    new_suggestion = Suggestion(
                        word=word,
                        russian_equivalent=russian_equivalent if russian_equivalent else None,
                        origin=origin if origin else None,
                        comment=comment if comment else None
                    )
                    db.session.add(new_suggestion)
                    db.session.commit()
                    print(f"Сохранено предложение: {word}")
            except Exception as e:
                print(f"Ошибка при сохранении в БД: {e}")

        return render_template('suggest.html', success=True)

    return render_template('suggest.html')


@app.route('/memes')
def memes():
    return render_template('memes.html')


@app.route('/download-excel')
def download_excel():
    return send_from_directory(app.config['UPLOAD_FOLDER'], 'dictionary.xlsx', as_attachment=True)


if __name__ == '__main__':
    with app.app_context():
        db.create_all()  # Дополнительная проверка создания таблиц
    app.run(host='0.0.0.0', port=8000)